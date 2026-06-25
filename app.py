import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
import re

# ══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="GSK Pending Order Report",
    page_icon="📦",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(bold=True, name="Arial", size=10)
BOLD_WHITE= Font(bold=True, color="FFFFFF", name="Arial", size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
_thin     = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
NO_BORDER = Border()

# Marketplace row colours (alternating)
NICK_FILL = {
    "shopee": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "tiktok": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "lazada": PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
}
NICK_FILL_ALT = {
    "shopee": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "tiktok": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "lazada": PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
}

# WMS status colours
WMS_FILL = {
    "FULFILLED":           PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "PENDING_FULFILLMENT": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "PARTIAL_ALLOCATED":   PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NONE_ALLOCATED":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CANCEL":              PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CANCELLED":           PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "RETURN":              PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CLOSE":               PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "NOT FOUND":           PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

# Pivot date-header colours
PIV_PAST   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
PIV_TODAY  = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
PIV_FUTURE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PIV_ROW_PAST   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PIV_ROW_TODAY  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PIV_ROW_FUTURE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
PIV_ROW_STATUS = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def nick_key(nick):
    """Return 'shopee' | 'tiktok' | 'lazada' from a nickname string."""
    n = str(nick or "").lower()
    for k in ("shopee", "tiktok", "lazada"):
        if k in n:
            return k
    return "shopee"


def row_fill(nick, row_idx):
    k = nick_key(nick)
    return NICK_FILL_ALT[k] if row_idx % 2 == 0 else NICK_FILL[k]


def add_one_day(date_str):
    """
    Parse a date string in any common format and return date+1 as YYYY-MM-DD.
    Handles: dd/mm/yyyy HH:MM:SS, yyyy-mm-dd HH:MM:SS, dd/mm/yyyy, yyyy-mm-dd, mm/dd/yyyy.
    """
    if not date_str:
        return ""
    s = str(date_str).strip()
    candidates = [s, s.split(" ")[0]] if " " in s else [s]
    for c in candidates:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y",
                    "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return (datetime.strptime(c, fmt) + timedelta(days=1)).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return ""


def fmt_pivot_date(d_str):
    """'2026-06-25' → '25-Jun-26'"""
    try:
        return datetime.strptime(d_str, "%Y-%m-%d").strftime("%d-%b-%y")
    except Exception:
        return d_str


def get_stock(raw_sku, inv_map):
    """Return integer stock for a SKU, or -1 if not found."""
    if not raw_sku:
        return -1
    base = raw_sku.split("x")[0]
    val  = inv_map.get(f"GSK_{base}", inv_map.get(base))
    if val is None:
        return -1
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return -1


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def write_header_row(ws, headers, col_widths=None):
    """Write a styled dark-blue header row and freeze pane below it."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CENTER; c.border = BORDER
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def write_data_row(ws, row_num, values, fill,
                   wms_col_idx=None, wms_val=None,
                   remarks_col_idx=None, remarks_val=None):
    """Write one data row with optional WMS and Remarks colour overrides."""
    for ci, v in enumerate(values, 1):
        c = ws.cell(row_num, ci, v)
        c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
        if wms_col_idx and ci == wms_col_idx and wms_val in WMS_FILL:
            c.fill = WMS_FILL[wms_val]
        elif remarks_col_idx and ci == remarks_col_idx:
            c.fill = RED_FILL if remarks_val == "CANCEL" else (
                     GREEN_FILL if remarks_val == "FULFILL" else fill)
        else:
            c.fill = fill


def clear_region(ws, max_row, max_col):
    """Blank out a rectangular region (1-based inclusive), removing all formatting."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill  = PatternFill()
            cell.font  = BODY_FONT
            cell.border = NO_BORDER


# ══════════════════════════════════════════════════════════════════════════════
#  PIVOT SHEET BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_pivot_sheet(ws, filtered_rows, headers, today_str):
    """
    Rebuild the PIVOT sheet from a list of 8-element rows.
    filtered_rows : list of lists  [Order Number, Invoice Number, Payment Status,
                                    Order Item Status, Ordered Date, Nickname, MP SLA, WMS]
    headers       : list of column names (same order as filtered_rows elements)
    today_str     : 'YYYY-MM-DD'
    """
    sla_idx    = headers.index("MP SLA")
    status_idx = headers.index("Order Item Status")

    pivot = defaultdict(lambda: defaultdict(int))
    dates_set, status_set = set(), set()
    for rv in filtered_rows:
        sla = rv[sla_idx] or ""
        st  = rv[status_idx] or ""
        if sla:
            pivot[st][sla] += 1
            dates_set.add(sla)
            status_set.add(st)

    col_dates = sorted(dates_set)
    statuses  = sorted(status_set)

    # Clear previous content generously
    old_r = max(ws.max_row or 1, len(statuses) + 5)
    old_c = max(ws.max_column or 1, len(col_dates) + 5)
    clear_region(ws, old_r, old_c)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 24

    # ── Header row ──────────────────────────────────────────────────────────
    c = ws.cell(1, 1, "Order Item Status")
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER

    for ci, d in enumerate(col_dates, 2):
        c = ws.cell(1, ci, fmt_pivot_date(d))
        c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 14
        if d < today_str:
            c.fill = PIV_PAST;   c.font = BOLD_WHITE
        elif d == today_str:
            c.fill = PIV_TODAY;  c.font = Font(bold=True, color="000000", name="Arial", size=10)
        else:
            c.fill = PIV_FUTURE; c.font = BOLD_WHITE

    gt_ci = len(col_dates) + 2
    c = ws.cell(1, gt_ci, "Grand Total")
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
    ws.column_dimensions[get_column_letter(gt_ci)].width = 14

    # ── Data rows ────────────────────────────────────────────────────────────
    col_grand = [0] * len(col_dates)
    for ri, st in enumerate(statuses, 2):
        ws.cell(ri, 1, st).fill   = PIV_ROW_STATUS
        ws.cell(ri, 1).font       = BODY_FONT
        ws.cell(ri, 1).border     = BORDER
        row_total = 0
        for ci, d in enumerate(col_dates, 2):
            val = pivot[st][d]
            col_grand[ci - 2] += val
            row_total += val
            c = ws.cell(ri, ci, val if val else "")
            c.alignment = CENTER; c.font = BODY_FONT; c.border = BORDER
            c.fill = PIV_ROW_PAST if d < today_str else (
                     PIV_ROW_TODAY if d == today_str else PIV_ROW_FUTURE)
        gt = ws.cell(ri, gt_ci, row_total)
        gt.alignment = CENTER; gt.font = BOLD_FONT; gt.border = BORDER

    # ── Grand Total row ──────────────────────────────────────────────────────
    gt_row     = len(statuses) + 2
    final_grand = sum(col_grand)
    ws.cell(gt_row, 1, "Grand Total").fill   = HDR_FILL
    ws.cell(gt_row, 1).font   = BOLD_WHITE
    ws.cell(gt_row, 1).border = BORDER
    for ci, v in enumerate(col_grand, 2):
        c = ws.cell(gt_row, ci, v if v else "")
        c.fill = HDR_FILL; c.font = BOLD_WHITE
        c.alignment = CENTER; c.border = BORDER
    c = ws.cell(gt_row, gt_ci, final_grand)
    c.fill = HDR_FILL; c.font = BOLD_WHITE
    c.alignment = CENTER; c.border = BORDER

    return final_grand


# ══════════════════════════════════════════════════════════════════════════════
#  CORE REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

DISPLAY_COLS  = ["Order Number", "Invoice Number", "Payment Status",
                 "Order Item Status", "Ordered Date", "Nickname", "MP SLA", "WMS"]
DISPLAY_WIDTHS = [20, 22, 16, 22, 22, 30, 14, 24]

PARTIAL_COLS   = DISPLAY_COLS + ["WMS Custom SKU", "Item Description",
                                  "Qty Ordered", "Qty Allocated", "WH Stock", "Remarks"]
PARTIAL_WIDTHS = DISPLAY_WIDTHS + [22, 35, 12, 14, 12, 12]

RMA_COLS   = ["Order Number", "Invoice Number", "Order Status",
              "Payment Status", "Ordered Date", "Nickname", "WMS"]
RMA_WIDTHS = [20, 22, 20, 16, 22, 30, 24]


def generate_report(tc_file, shopee_file, wms_files, inv_file):
    """
    Build the 7-sheet GSK Pending Order Report workbook.

    Parameters
    ----------
    tc_file     : file-like  — TC Order Report CSV
    shopee_file : file-like  — Shopee order XLSX (Estimated Ship Out Date)
    wms_files   : list[file] — one or more WMS XLSX files (auto-concatenated)
    inv_file    : file-like  — InventoryWarehouse Report XLSX

    Returns
    -------
    (wb, summary_dict)
    """

    # ── Load TC ──────────────────────────────────────────────────────────────
    tc = pd.read_csv(tc_file, dtype=str)
    tc = tc.apply(lambda col: col.str.replace("\r", "", regex=False).str.strip())
    tc.replace("nan", "", inplace=True)

    # ── Load Shopee SLA map ───────────────────────────────────────────────────
    sh = pd.read_excel(shopee_file, dtype=str)
    sh["Order ID"] = sh["Order ID"].astype(str).str.strip()
    sh["Estimated Ship Out Date"] = sh["Estimated Ship Out Date"].astype(str).str.strip()
    shopee_sla = sh.drop_duplicates("Order ID").set_index("Order ID")["Estimated Ship Out Date"].to_dict()

    # ── Load WMS (one or many files) ─────────────────────────────────────────
    wms_dfs = []
    for f in wms_files:
        try:
            wms_dfs.append(pd.read_excel(f, dtype=str))
        except Exception as e:
            st.warning(f"Skipped WMS file ({e})")
    wms = pd.concat(wms_dfs, ignore_index=True) if wms_dfs else pd.DataFrame(
        columns=["client_order_id", "status_so"])
    wms["client_order_id"] = wms["client_order_id"].astype(str).str.strip()
    wms["status_so"]       = wms["status_so"].astype(str).str.strip()
    wms_status_map = wms.drop_duplicates("client_order_id").set_index("client_order_id")["status_so"].to_dict()

    # ── Load Inventory ────────────────────────────────────────────────────────
    inv = pd.read_excel(inv_file, dtype=str)
    inv_map = {}
    if "client_code_item" in inv.columns:
        # InventoryWarehouse format — stock per SKU line
        inv["quantity_available"] = inv["quantity_available"].apply(
            lambda x: "0" if str(x).strip() in ("nan", "", "None") else str(x).strip())
        inv_map = inv.set_index("client_code_item")["quantity_available"].to_dict()
    elif "code_item" in inv.columns:
        # InventoryZone format — aggregate across zones
        inv["quantity_available"] = pd.to_numeric(inv["quantity_available"], errors="coerce").fillna(0)
        inv_map = inv.groupby("code_item")["quantity_available"].sum().to_dict()

    # ── SLA helper ────────────────────────────────────────────────────────────
    def get_mp_sla(row):
        nick = str(row["nickname"]).lower()
        if "shopee" in nick:
            val = (shopee_sla.get(str(row["order_id"]).strip()) or
                   shopee_sla.get(str(row["order_number"]).strip()) or "")
            return val[:10] if val and val != "nan" else ""
        if "tiktok" in nick or "lazada" in nick:
            return add_one_day(row["ordered_date"])
        return ""

    # ── Filter pending orders ─────────────────────────────────────────────────
    PENDING_STATUSES = {"ACCEPTED/PICKED", "NEW", "READY TO SHIP"}
    mask = (
        tc["order_item_status"].isin(PENDING_STATUSES) &
        tc["nickname"].str.lower().str.contains("tiktok|lazada|shopee", na=False) &
        (
            (tc["payment_status"] == "COMPLETED") |
            ((tc["payment_status"] == "PENDING") & tc["payment_methods"].str.contains("COD", na=False))
        )
    )
    filtered = tc[mask].copy()
    filtered["MP SLA"] = filtered.apply(get_mp_sla, axis=1)
    filtered["WMS"]    = filtered["invoice_number"].map(wms_status_map).fillna("NOT FOUND")
    filtered["_stock"] = filtered["custom_sku"].apply(lambda x: get_stock(x, inv_map))

    # ── Sub-groups ───────────────────────────────────────────────────────────
    is_partial    = filtered["WMS"].isin({"PARTIAL_ALLOCATED", "NONE_ALLOCATED"})
    is_zero_stock = is_partial & (filtered["_stock"] == 0)
    is_cancel     = filtered["WMS"].isin({"CANCEL", "CANCELLED"})

    # FILTERED DATA: exclude NOT FOUND, CANCEL/CANCELLED, and 0-stock partial
    df_filtered = filtered[
        ~filtered["WMS"].isin({"NOT FOUND"}) &
        ~is_cancel &
        ~is_zero_stock
    ].copy()

    df_not_pushed = filtered[filtered["WMS"] == "NOT FOUND"].copy()
    df_partial    = filtered[is_partial].copy()
    df_close_nf   = filtered[filtered["WMS"].isin({"CLOSE", "NOT FOUND"})].copy()

    # RMA: cancelled orders that are FULFILLED in WMS
    rma_raw = tc[tc["order_status"].isin({"CANCELLED", "CANCEL REQUESTED"})].copy()
    rma_raw["WMS"] = rma_raw["invoice_number"].map(wms_status_map).fillna("NOT FOUND")
    df_rma = rma_raw[rma_raw["WMS"] == "FULFILLED"].drop_duplicates("invoice_number")

    # ── Build workbook ────────────────────────────────────────────────────────
    today_str = datetime.today().strftime("%Y-%m-%d")
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: FILTERED DATA ────────────────────────────────────────────────
    ws1 = wb.create_sheet("FILTERED DATA")
    ws1.sheet_properties.tabColor = "4472C4"
    write_header_row(ws1, DISPLAY_COLS, DISPLAY_WIDTHS)
    fd_rows = []
    for i, (_, r) in enumerate(df_filtered.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                r["order_item_status"], r["ordered_date"], r["nickname"],
                r["MP SLA"], r["WMS"]]
        fd_rows.append(vals)
        write_data_row(ws1, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=8, wms_val=r["WMS"])

    # ── Sheet 2: PIVOT ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("PIVOT")
    ws2.sheet_properties.tabColor = "4472C4"
    build_pivot_sheet(ws2, fd_rows, DISPLAY_COLS, today_str)

    # ── Sheet 3: NOT PUSHED ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("NOT PUSHED")
    ws3.sheet_properties.tabColor = "FF0000"
    write_header_row(ws3, DISPLAY_COLS, DISPLAY_WIDTHS)
    for i, (_, r) in enumerate(df_not_pushed.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                r["order_item_status"], r["ordered_date"], r["nickname"],
                r["MP SLA"], r["WMS"]]
        write_data_row(ws3, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=8, wms_val=r["WMS"])

    # ── Sheet 4: PARTIAL & NON ALLOCATED ─────────────────────────────────────
    ws4 = wb.create_sheet("PARTIAL & NON ALLOCATED")
    ws4.sheet_properties.tabColor = "ED7D31"
    write_header_row(ws4, PARTIAL_COLS, PARTIAL_WIDTHS)

    # Build WMS detail lookup for PARTIAL rows (item_descr, qty ordered/allocated)
    wms_detail_idx = {}
    if "client_order_id" in wms.columns:
        for _, wr in wms.iterrows():
            oid = str(wr.get("client_order_id", "")).strip()
            if oid and oid not in wms_detail_idx:
                wms_detail_idx[oid] = wr

    for i, (_, r) in enumerate(df_partial.iterrows(), 1):
        inv_num  = str(r["invoice_number"]).strip()
        raw_sku  = str(r["custom_sku"]).strip()
        sku_base = raw_sku.split("x")[0] if raw_sku else ""
        wms_sku  = f"GSK_{sku_base}" if sku_base else ""
        stock    = r["_stock"]
        remarks  = "CANCEL" if stock == 0 else ("FULFILL" if stock > 0 else "")

        wd = wms_detail_idx.get(inv_num)
        item_descr  = str(wd.get("item_descr", ""))        if wd is not None else ""
        qty_ordered = str(wd.get("quantity", ""))          if wd is not None else ""
        qty_alloc   = str(wd.get("quantity_allocated", "")) if wd is not None else ""

        vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                r["order_item_status"], r["ordered_date"], r["nickname"],
                r["MP SLA"], r["WMS"],
                wms_sku, item_descr, qty_ordered, qty_alloc,
                stock if stock >= 0 else "", remarks]

        rf = row_fill(r["nickname"], i)
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(i + 1, ci, v)
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
            if ci == 8 and r["WMS"] in WMS_FILL:
                c.fill = WMS_FILL[r["WMS"]]
            elif ci == 14:  # Remarks column
                c.fill = RED_FILL if remarks == "CANCEL" else (GREEN_FILL if remarks == "FULFILL" else rf)
            else:
                c.fill = rf

    # ── Sheet 5: CLOSE & NOT FOUND ────────────────────────────────────────────
    ws5 = wb.create_sheet("CLOSE & NOT FOUND")
    ws5.sheet_properties.tabColor = "808080"
    write_header_row(ws5, DISPLAY_COLS, DISPLAY_WIDTHS)
    for i, (_, r) in enumerate(df_close_nf.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                r["order_item_status"], r["ordered_date"], r["nickname"],
                r["MP SLA"], r["WMS"]]
        write_data_row(ws5, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=8, wms_val=r["WMS"])

    # ── Sheet 6: ORIGINAL (All Data) ──────────────────────────────────────────
    ws6 = wb.create_sheet("ORIGINAL (All Data)")
    ws6.sheet_properties.tabColor = "1F1F1F"
    write_header_row(ws6, list(tc.columns))
    for ri, (_, r) in enumerate(tc.iterrows(), 2):
        for ci, col in enumerate(tc.columns, 1):
            c = ws6.cell(ri, ci, r[col])
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER

    # ── Sheet 7: RMA ──────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("RMA")
    ws7.sheet_properties.tabColor = "C00000"
    write_header_row(ws7, RMA_COLS, RMA_WIDTHS)
    for i, (_, r) in enumerate(df_rma.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["order_status"],
                r["payment_status"], r["ordered_date"], r["nickname"], r["WMS"]]
        write_data_row(ws7, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=7, wms_val=r["WMS"])

    summary = {
        "FILTERED DATA":           len(df_filtered),
        "NOT PUSHED":              len(df_not_pushed),
        "PARTIAL & NON ALLOCATED": len(df_partial),
        "CLOSE & NOT FOUND":       len(df_close_nf),
        "ORIGINAL":                len(tc),
        "RMA":                     len(df_rma),
    }
    return wb, summary


# ══════════════════════════════════════════════════════════════════════════════
#  NOT-PUSHED UPDATER
# ══════════════════════════════════════════════════════════════════════════════

def update_not_pushed(report_bytes, invoice_list, wms_status="PENDING_FULFILLMENT"):
    """
    Given an existing report (bytes) and a list of invoice numbers that have
    now been pushed to WMS:
      1. Read those rows from NOT PUSHED
      2. Set their WMS column to wms_status
      3. Append them to FILTERED DATA
      4. Remove them from NOT PUSHED and CLOSE & NOT FOUND
      5. Rebuild PIVOT
      PARTIAL, ORIGINAL, RMA are left completely untouched.

    Returns (updated_bytes, matched_count, not_found_set)
    """
    wb        = load_workbook(BytesIO(report_bytes))
    today_str = datetime.today().strftime("%Y-%m-%d")
    invoices  = {str(i).strip().upper() for i in invoice_list if str(i).strip()}

    NUM_COLS    = 8   # all standard sheets have exactly 8 data columns
    WMS_COL    = 7   # 0-based index of WMS in an 8-element row
    NICK_COL   = 5   # 0-based index of Nickname

    def read_rows(ws):
        """Return (headers_list, data_rows_list) reading exactly NUM_COLS columns."""
        headers = [ws.cell(1, c).value for c in range(1, NUM_COLS + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            rv = [ws.cell(r, c).value for c in range(1, NUM_COLS + 1)]
            if any(v is not None for v in rv):
                rows.append(rv)
        return headers, rows

    def rewrite_sheet(ws, rows):
        """
        Clear all data rows (with a safe ceiling) and write new rows.
        Header row (row 1) is preserved.
        """
        ceiling = max(ws.max_row or 1, len(rows) + 20)
        for r in range(2, ceiling + 1):
            for c in range(1, NUM_COLS + 1):
                cell = ws.cell(r, c)
                cell.value  = None
                cell.fill   = PatternFill()
                cell.font   = BODY_FONT
                cell.border = NO_BORDER
        for i, rv in enumerate(rows, 1):
            nick    = rv[NICK_COL] if len(rv) > NICK_COL else ""
            wms_val = rv[WMS_COL]  if len(rv) > WMS_COL  else None
            rf      = row_fill(nick, i)
            write_data_row(ws, i + 1, rv, rf, wms_col_idx=8, wms_val=wms_val)

    # 1. Read FILTERED DATA
    ws_fd              = wb["FILTERED DATA"]
    fd_headers, fd_rows = read_rows(ws_fd)

    # 2. Partition NOT PUSHED into matched vs kept
    ws_np              = wb["NOT PUSHED"]
    np_headers, np_rows = read_rows(ws_np)
    inv_col            = np_headers.index("Invoice Number")

    matched, kept_np, matched_inv = [], [], set()
    for rv in np_rows:
        key = str(rv[inv_col] or "").strip().upper()
        if key in invoices:
            updated = list(rv[:WMS_COL]) + [wms_status]  # replace WMS column
            matched.append(updated)
            matched_inv.add(key)
        else:
            kept_np.append(rv)

    # 3. Remove matched from CLOSE & NOT FOUND
    ws_cnf               = wb["CLOSE & NOT FOUND"]
    cnf_headers, cnf_rows = read_rows(ws_cnf)
    cnf_inv_col          = cnf_headers.index("Invoice Number")
    kept_cnf = [rv for rv in cnf_rows
                if str(rv[cnf_inv_col] or "").strip().upper() not in invoices]

    # 4. Merge matched into FILTERED DATA
    final_fd = fd_rows + matched

    # 5. Rewrite the three changed sheets
    rewrite_sheet(ws_fd,  final_fd)
    rewrite_sheet(ws_np,  kept_np)
    rewrite_sheet(ws_cnf, kept_cnf)

    # 6. Rebuild PIVOT only — PARTIAL, ORIGINAL, RMA untouched
    build_pivot_sheet(wb["PIVOT"], final_fd, fd_headers, today_str)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), len(matched), invoices - matched_inv


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.title("📦 GSK Pending Order Report")
st.markdown("---")

tab1, tab2 = st.tabs(["🗂️ Generate Report", "✏️ Update NOT PUSHED Orders"])


# ─────────────────────────────────────────────────────────────────────────────
#  TAB 1 — GENERATE REPORT
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    st.subheader("Upload daily source files")

    col_a, col_b = st.columns(2)
    with col_a:
        tc_file     = st.file_uploader("TC Order Report (.csv)",    type=["csv"],  key="tc")
        shopee_file = st.file_uploader("Shopee Order File (.xlsx)", type=["xlsx"], key="shopee")
    with col_b:
        wms_files   = st.file_uploader("WMS File(s) (.xlsx) — upload one or more",
                                        type=["xlsx"], key="wms", accept_multiple_files=True)
        inv_file    = st.file_uploader("Inventory Report (.xlsx)",  type=["xlsx"], key="inv")

    if st.button("🚀 Generate Report", type="primary", use_container_width=True):
        if not all([tc_file, shopee_file, wms_files, inv_file]):
            st.error("Please upload all four file types before generating.")
        else:
            with st.spinner("Processing files…"):
                try:
                    wb, summary = generate_report(tc_file, shopee_file, wms_files, inv_file)

                    buf = BytesIO()
                    wb.save(buf)
                    report_bytes = buf.getvalue()

                    fname = f"GSK_Pending_Order_Report_{datetime.today().strftime('%Y%m%d')}.xlsx"
                    st.session_state["report_bytes"]   = report_bytes
                    st.session_state["report_filename"] = fname

                    st.success("✅ Report generated successfully!")

                    # Summary metrics
                    st.markdown("#### Summary")
                    icons = {"FILTERED DATA": "🔵", "NOT PUSHED": "🔴",
                             "PARTIAL & NON ALLOCATED": "🟠", "CLOSE & NOT FOUND": "⚫",
                             "ORIGINAL": "⬛", "RMA": "🔴"}
                    mcols = st.columns(len(summary))
                    for col, (k, v) in zip(mcols, summary.items()):
                        col.metric(f"{icons.get(k, '')} {k}", v)

                    st.download_button(
                        "⬇️ Download Report", data=report_bytes, file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                    if summary["NOT PUSHED"] > 0:
                        st.info(
                            f"💡 **{summary['NOT PUSHED']} order(s) NOT PUSHED** — "
                            "once you confirm which are now in WMS, go to the "
                            "**Update NOT PUSHED Orders** tab and paste the invoice numbers."
                        )

                except Exception as e:
                    st.error(f"Error: {e}")
                    st.exception(e)


# ─────────────────────────────────────────────────────────────────────────────
#  TAB 2 — UPDATE NOT PUSHED ORDERS
# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    st.subheader("Update orders confirmed as pushed to WMS")
    st.markdown(
        "The report generated in Tab 1 is loaded automatically. "
        "Or upload an existing report file below."
    )

    existing = st.file_uploader(
        "Upload existing report (.xlsx)  *(optional if already generated above)*",
        type=["xlsx"], key="existing_report",
    )
    if existing:
        st.session_state["report_bytes"]   = existing.read()
        st.session_state["report_filename"] = existing.name

    has_report = bool(st.session_state.get("report_bytes"))
    if has_report:
        st.success(f"📄 Loaded: **{st.session_state.get('report_filename', 'report.xlsx')}**")
    else:
        st.warning("No report loaded. Generate one in Tab 1 or upload an existing file above.")

    st.markdown("---")
    st.markdown(
        "#### Paste invoice numbers that are now in WMS\n"
        "One per line, or comma-separated — e.g. `GSK00352711, GSK00352812`"
    )

    invoice_text = st.text_area(
        "Invoice Numbers", height=160, key="invoice_input",
        placeholder="GSK00352711\nGSK00352812\nGSK00353001",
    )

    wms_status = st.selectbox(
        "WMS status to assign to these orders",
        options=["PENDING_FULFILLMENT", "FULFILLED", "PARTIAL_ALLOCATED"],
        index=0, key="wms_status_select",
    )

    if st.button("✅ Update Report", type="primary",
                 use_container_width=True, disabled=not has_report):
        raw = invoice_text.strip()
        if not raw:
            st.error("Please enter at least one invoice number.")
        else:
            invoices = [p.strip() for p in re.split(r"[\n,]+", raw) if p.strip()]
            with st.spinner(f"Updating {len(invoices)} invoice(s)…"):
                try:
                    updated_bytes, matched, not_found = update_not_pushed(
                        st.session_state["report_bytes"], invoices, wms_status
                    )
                    st.session_state["report_bytes"] = updated_bytes

                    if matched:
                        st.success(
                            f"✅ **{matched} invoice(s)** moved from NOT PUSHED → FILTERED DATA "
                            f"with WMS status `{wms_status}`. PIVOT rebuilt."
                        )
                    else:
                        st.warning("None of the invoice numbers were found in the NOT PUSHED sheet.")

                    if not_found:
                        st.warning(
                            f"⚠️ Not found in NOT PUSHED: {', '.join(sorted(not_found))}"
                        )

                    st.download_button(
                        "⬇️ Download Updated Report",
                        data=updated_bytes,
                        file_name=st.session_state.get("report_filename", "GSK_Report_Updated.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"Error: {e}")
                    st.exception(e)
