import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
import re

# ══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="GSK Pending Order Report",
    page_icon="📦",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(bold=True, name="Arial", size=10)
BOLD_WHITE= Font(bold=True, color="FFFFFF", name="Arial", size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
_thin     = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
NO_BORDER = Border()

# Marketplace row colours (alternating)
NICK_FILL = {
    "shopee": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "tiktok": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "lazada": PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
}
NICK_FILL_ALT = {
    "shopee": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "tiktok": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "lazada": PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
}

# WMS status colours
WMS_FILL = {
    "FULFILLED":           PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "PENDING_FULFILLMENT": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "PARTIAL_ALLOCATED":   PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NONE_ALLOCATED":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CANCEL":              PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CANCELLED":           PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "RETURN":              PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "CLOSE":               PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "NOT FOUND":           PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

# Pivot date-header colours
PIV_PAST   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
PIV_TODAY  = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
PIV_FUTURE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PIV_ROW_PAST   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PIV_ROW_TODAY  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PIV_ROW_FUTURE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
PIV_ROW_STATUS = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def nick_key(nick):
    """Return 'shopee' | 'tiktok' | 'lazada' from a nickname string."""
    n = str(nick or "").lower()
    for k in ("shopee", "tiktok", "lazada"):
        if k in n:
            return k
    return "shopee"


def row_fill(nick, row_idx):
    k = nick_key(nick)
    return NICK_FILL_ALT[k] if row_idx % 2 == 0 else NICK_FILL[k]


def add_one_day(date_str):
    """
    Parse a date string in any common format and return date+1 as YYYY-MM-DD.
    Handles: dd/mm/yyyy HH:MM:SS, yyyy-mm-dd HH:MM:SS, dd/mm/yyyy, yyyy-mm-dd,
             mm/dd/yyyy, dd Mon yyyy HH:MM (Lazada format e.g. '03 Jul 2026 08:40').
    """
    if not date_str:
        return ""
    s = str(date_str).strip()
    candidates = [s, s.split(" ")[0]] if " " in s else [s]
    for c in candidates:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y",
                    "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                    "%d %b %Y %H:%M", "%d %b %Y"):
            try:
                return (datetime.strptime(c, fmt) + timedelta(days=1)).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return ""


def fmt_pivot_date(d_str):
    """'2026-06-25' → '25-Jun-26'"""
    try:
        return datetime.strptime(d_str, "%Y-%m-%d").strftime("%d-%b-%y")
    except Exception:
        return d_str


def get_stock(raw_sku, inv_map):
    """Return integer stock for a SKU, or -1 if not found."""
    if not raw_sku:
        return -1
    base = raw_sku.split("x")[0]
    val  = inv_map.get(f"GSK_{base}", inv_map.get(base))
    if val is None:
        return -1
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return -1


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def write_header_row(ws, headers, col_widths=None):
    """Write a styled dark-blue header row and freeze pane below it."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CENTER; c.border = BORDER
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def write_data_row(ws, row_num, values, fill,
                   wms_col_idx=None, wms_val=None,
                   remarks_col_idx=None, remarks_val=None):
    """Write one data row with optional WMS and Remarks colour overrides."""
    for ci, v in enumerate(values, 1):
        c = ws.cell(row_num, ci, v)
        c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
        if wms_col_idx and ci == wms_col_idx and wms_val in WMS_FILL:
            c.fill = WMS_FILL[wms_val]
        elif remarks_col_idx and ci == remarks_col_idx:
            c.fill = RED_FILL if remarks_val == "CANCEL" else (
                     GREEN_FILL if remarks_val == "FULFILL" else fill)
        else:
            c.fill = fill


def clear_region(ws, max_row, max_col):
    """Blank out a rectangular region (1-based inclusive), removing all formatting."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill  = PatternFill()
            cell.font  = BODY_FONT
            cell.border = NO_BORDER


# ══════════════════════════════════════════════════════════════════════════════
#  PIVOT SHEET BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_pivot_sheet(ws, filtered_rows, headers, today_str):
    """
    Rebuild the PIVOT sheet from a list of 8-element rows.
    filtered_rows : list of lists  [Order Number, Invoice Number, Payment Status,
                                    Order Item Status, Ordered Date, Nickname, MP SLA, WMS]
    headers       : list of column names (same order as filtered_rows elements)
    today_str     : 'YYYY-MM-DD'
    """
    sla_idx    = headers.index("MP SLA")
    status_idx = headers.index("Order Item Status")

    pivot = defaultdict(lambda: defaultdict(int))
    dates_set, status_set = set(), set()
    for rv in filtered_rows:
        sla = rv[sla_idx] or ""
        st  = rv[status_idx] or ""
        if sla:
            pivot[st][sla] += 1
            dates_set.add(sla)
            status_set.add(st)

    col_dates = sorted(dates_set)
    statuses  = sorted(status_set)

    # Clear previous content generously
    old_r = max(ws.max_row or 1, len(statuses) + 5)
    old_c = max(ws.max_column or 1, len(col_dates) + 5)
    clear_region(ws, old_r, old_c)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 24

    # ── Header row ──────────────────────────────────────────────────────────
    c = ws.cell(1, 1, "Order Item Status")
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER

    for ci, d in enumerate(col_dates, 2):
        c = ws.cell(1, ci, fmt_pivot_date(d))
        c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 14
        if d < today_str:
            c.fill = PIV_PAST;   c.font = BOLD_WHITE
        elif d == today_str:
            c.fill = PIV_TODAY;  c.font = Font(bold=True, color="000000", name="Arial", size=10)
        else:
            c.fill = PIV_FUTURE; c.font = BOLD_WHITE

    gt_ci = len(col_dates) + 2
    c = ws.cell(1, gt_ci, "Grand Total")
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
    ws.column_dimensions[get_column_letter(gt_ci)].width = 14

    # ── Data rows ────────────────────────────────────────────────────────────
    col_grand = [0] * len(col_dates)
    for ri, st in enumerate(statuses, 2):
        ws.cell(ri, 1, st).fill   = PIV_ROW_STATUS
        ws.cell(ri, 1).font       = BODY_FONT
        ws.cell(ri, 1).border     = BORDER
        row_total = 0
        for ci, d in enumerate(col_dates, 2):
            val = pivot[st][d]
            col_grand[ci - 2] += val
            row_total += val
            c = ws.cell(ri, ci, val if val else "")
            c.alignment = CENTER; c.font = BODY_FONT; c.border = BORDER
            c.fill = PIV_ROW_PAST if d < today_str else (
                     PIV_ROW_TODAY if d == today_str else PIV_ROW_FUTURE)
        gt = ws.cell(ri, gt_ci, row_total)
        gt.alignment = CENTER; gt.font = BOLD_FONT; gt.border = BORDER

    # ── Grand Total row ──────────────────────────────────────────────────────
    gt_row     = len(statuses) + 2
    final_grand = sum(col_grand)
    ws.cell(gt_row, 1, "Grand Total").fill   = HDR_FILL
    ws.cell(gt_row, 1).font   = BOLD_WHITE
    ws.cell(gt_row, 1).border = BORDER
    for ci, v in enumerate(col_grand, 2):
        c = ws.cell(gt_row, ci, v if v else "")
        c.fill = HDR_FILL; c.font = BOLD_WHITE
        c.alignment = CENTER; c.border = BORDER
    c = ws.cell(gt_row, gt_ci, final_grand)
    c.fill = HDR_FILL; c.font = BOLD_WHITE
    c.alignment = CENTER; c.border = BORDER

    return final_grand


# ══════════════════════════════════════════════════════════════════════════════
#  CORE REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

DISPLAY_COLS  = ["Order Number", "Invoice Number", "Payment Status",
                 "Order Item Status", "Ordered Date", "Nickname", "MP SLA", "WMS"]
DISPLAY_WIDTHS = [20, 22, 16, 22, 22, 30, 14, 24]

# PARTIAL sheet: no Order Number — starts with Invoice Number
PARTIAL_COLS   = ["Order Number", "Invoice Number", "Payment Status", "Order Item Status",
                  "Ordered Date", "Nickname", "MP SLA", "WMS",
                  "WMS Custom SKU", "Item Description",
                  "Qty Ordered", "Qty Allocated", "WH Stock", "Remarks"]
PARTIAL_WIDTHS = [20, 22, 16, 22, 22, 30, 14, 24, 28, 35, 12, 14, 12, 12]
PARTIAL_ORDER_COLS = 8   # first 8 cols are order-level (blanked on sub-rows)
PARTIAL_WMS_COL   = 8   # 1-based index of WMS column in PARTIAL sheet

RMA_COLS   = ["Order Number", "Invoice Number", "Order Status",
              "Payment Status", "Ordered Date", "Nickname", "WMS"]
RMA_WIDTHS = [20, 22, 20, 16, 22, 30, 24]


def generate_report(tc_file, shopee_file, wms_files, inv_file,
                    tiktok_files=None, lazada_file=None):
    """
    Build the GSK Pending Order Report workbook.

    Parameters
    ----------
    tc_file      : file-like  — TC Order Report CSV
    shopee_file  : file-like  — Shopee order XLSX (used for Estimated Ship Out Date
                   AND cross-checked against TC for the NOT IN TC sheet)
    wms_files    : list[file] — one or more WMS XLSX files (auto-concatenated)
    inv_file     : file-like  — InventoryWarehouse Report XLSX
    tiktok_files : list[file] — optional TikTok order XLSX files (any number of stores)
    lazada_file  : file-like  — optional Lazada order XLSX

    Returns
    -------
    (wb, summary_dict)
    """
    tiktok_files = tiktok_files or []

    # ── Load TC ──────────────────────────────────────────────────────────────
    tc = pd.read_csv(BytesIO(tc_file.read()), dtype=str)
    tc = tc.apply(lambda col: col.str.replace("\r", "", regex=False).str.strip())
    tc.replace("nan", "", inplace=True)

    # ── Load Shopee SLA map ───────────────────────────────────────────────────
    sh = pd.read_excel(BytesIO(shopee_file.read()), dtype=str)
    sh["Order ID"] = sh["Order ID"].astype(str).str.strip()
    sh["Estimated Ship Out Date"] = sh["Estimated Ship Out Date"].astype(str).str.strip()
    shopee_sla = sh.drop_duplicates("Order ID").set_index("Order ID")["Estimated Ship Out Date"].to_dict()

    # ── Load WMS (one or many files) ─────────────────────────────────────────
    wms_bytes_list = []
    for f in wms_files:
        try:
            wms_bytes_list.append(f.read())  # read bytes immediately — UploadedFile exhausts after first read
        except Exception as e:
            st.warning(f"Skipped WMS file ({e})")

    wms_dfs = []
    for b in wms_bytes_list:
        try:
            wms_dfs.append(pd.read_excel(BytesIO(b), dtype=str))
        except Exception as e:
            st.warning(f"Could not parse WMS file ({e})")

    wms = pd.concat(wms_dfs, ignore_index=True) if wms_dfs else pd.DataFrame(
        columns=["client_order_id", "status_so"])
    wms.columns = wms.columns.str.strip()
    wms["client_order_id"] = wms["client_order_id"].astype(str).str.strip()
    wms["status_so"]       = wms["status_so"].astype(str).str.strip()
    wms_status_map = wms.drop_duplicates("client_order_id").set_index("client_order_id")["status_so"].to_dict()

    # ── Load Inventory ────────────────────────────────────────────────────────
    inv = pd.read_excel(BytesIO(inv_file.read()), dtype=str)
    inv_map = {}
    if "client_code_item" in inv.columns:
        # InventoryWarehouse format — stock per SKU line
        inv["quantity_available"] = inv["quantity_available"].apply(
            lambda x: "0" if str(x).strip() in ("nan", "", "None") else str(x).strip())
        inv_map = inv.set_index("client_code_item")["quantity_available"].to_dict()
    elif "code_item" in inv.columns:
        # InventoryZone format — aggregate across zones
        inv["quantity_available"] = pd.to_numeric(inv["quantity_available"], errors="coerce").fillna(0)
        inv_map = inv.groupby("code_item")["quantity_available"].sum().to_dict()

    # ── SLA helper ────────────────────────────────────────────────────────────
    def get_mp_sla(row):
        nick = str(row["nickname"]).lower()
        if "shopee" in nick:
            val = (shopee_sla.get(str(row["order_id"]).strip()) or
                   shopee_sla.get(str(row["order_number"]).strip()) or "")
            return val[:10] if val and val != "nan" else ""
        if "tiktok" in nick or "lazada" in nick:
            return add_one_day(row["ordered_date"])
        return ""

    # ── Filter pending orders ─────────────────────────────────────────────────
    PENDING_STATUSES = {"ACCEPTED/PICKED", "NEW", "READY TO SHIP"}
    mask = (
        tc["order_item_status"].isin(PENDING_STATUSES) &
        tc["nickname"].str.lower().str.contains("tiktok|lazada|shopee", na=False) &
        (
            (tc["payment_status"] == "COMPLETED") |
            ((tc["payment_status"] == "PENDING") & tc["payment_methods"].str.contains("COD", na=False))
        )
    )
    filtered = tc[mask].copy()
    filtered["MP SLA"] = filtered.apply(get_mp_sla, axis=1)
    filtered["WMS"]    = filtered["invoice_number"].map(wms_status_map).fillna("NOT FOUND")
    filtered["_stock"] = filtered["custom_sku"].apply(lambda x: get_stock(x, inv_map))

    # ── Sub-groups ───────────────────────────────────────────────────────────
    is_partial    = filtered["WMS"].isin({"PARTIAL_ALLOCATED", "NONE_ALLOCATED"})
    is_zero_stock = is_partial & (filtered["_stock"] == 0)
    is_cancel     = filtered["WMS"].isin({"CANCEL", "CANCELLED"})

    # Also catch partial/none orders that exist in WMS but didn't pass the TC filter
    # (e.g. offline/B2B orders with blank order_item_status or non-marketplace nickname)
    wms_partial_invs = set(
        wms.loc[wms["status_so"].isin({"PARTIAL_ALLOCATED", "NONE_ALLOCATED"}), "client_order_id"]
    )
    tc_partial_extra = pd.DataFrame()
    # Get their TC rows (all rows, no filter) and mark them
    tc_extra_rows = tc[
        tc["invoice_number"].isin(wms_partial_invs) &
        ~tc["invoice_number"].isin(filtered[is_partial]["invoice_number"])
    ].copy()
    if len(tc_extra_rows):
        tc_extra_rows["MP SLA"] = tc_extra_rows.apply(get_mp_sla, axis=1)
        tc_extra_rows["WMS"]    = tc_extra_rows["invoice_number"].map(wms_status_map).fillna("NOT FOUND")
        tc_extra_rows["_stock"] = tc_extra_rows["custom_sku"].apply(lambda x: get_stock(x, inv_map))
        tc_partial_extra = tc_extra_rows

    # FILTERED DATA: exclude NOT FOUND, CANCEL/CANCELLED, and 0-stock partial
    df_filtered = filtered[
        ~filtered["WMS"].isin({"NOT FOUND"}) &
        ~is_cancel &
        ~is_zero_stock
    ].copy()

    df_not_pushed = filtered[filtered["WMS"] == "NOT FOUND"].copy()
    df_partial    = filtered[is_partial].copy()
    # Merge in any WMS-partial orders not caught by the main filter
    if len(tc_partial_extra):
        df_partial = pd.concat([df_partial, tc_partial_extra], ignore_index=True)
    df_close_nf   = filtered[filtered["WMS"].isin({"CLOSE", "NOT FOUND"})].copy()

    # RMA: cancelled orders that are FULFILLED in WMS
    rma_raw = tc[tc["order_status"].isin({"CANCELLED", "CANCEL REQUESTED"})].copy()
    rma_raw["WMS"] = rma_raw["invoice_number"].map(wms_status_map).fillna("NOT FOUND")
    df_rma = rma_raw[rma_raw["WMS"] == "FULFILLED"].drop_duplicates("invoice_number")

    # ── Build workbook ────────────────────────────────────────────────────────
    today_str = datetime.today().strftime("%Y-%m-%d")
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: FILTERED DATA ────────────────────────────────────────────────
    ws1 = wb.create_sheet("FILTERED DATA")
    ws1.sheet_properties.tabColor = "4472C4"
    write_header_row(ws1, DISPLAY_COLS, DISPLAY_WIDTHS)
    fd_rows = []
    for i, (_, r) in enumerate(df_filtered.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                r["order_item_status"], r["ordered_date"], r["nickname"],
                r["MP SLA"], r["WMS"]]
        fd_rows.append(vals)
        write_data_row(ws1, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=8, wms_val=r["WMS"])

    # ── Sheet 2: PIVOT ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("PIVOT")
    ws2.sheet_properties.tabColor = "4472C4"
    build_pivot_sheet(ws2, fd_rows, DISPLAY_COLS, today_str)

    # ── Sheet 3: NOT PUSHED TO WMS ────────────────────────────────────────────
    ws3 = wb.create_sheet("NOT PUSHED TO WMS")
    ws3.sheet_properties.tabColor = "FF0000"
    write_header_row(ws3, DISPLAY_COLS, DISPLAY_WIDTHS)
    for i, (_, r) in enumerate(df_not_pushed.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                r["order_item_status"], r["ordered_date"], r["nickname"],
                r["MP SLA"], r["WMS"]]
        write_data_row(ws3, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=8, wms_val=r["WMS"])

    # ── Sheet 4: PARTIAL & NON ALLOCATED ─────────────────────────────────────
    ws4 = wb.create_sheet("PARTIAL & NON ALLOCATED")
    ws4.sheet_properties.tabColor = "ED7D31"
    write_header_row(ws4, PARTIAL_COLS, PARTIAL_WIDTHS)

    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Build WMS detail lookup: ALL lines per invoice (multi-SKU support)
    # Strip all column names first to handle any whitespace/BOM issues
    wms.columns = wms.columns.str.strip()
    wms_lines_by_inv = defaultdict(list)
    order_id_col = "client_order_id" if "client_order_id" in wms.columns else None
    if order_id_col:
        for row_dict in wms.to_dict("records"):
            oid = str(row_dict.get(order_id_col, "")).strip()
            if oid:
                wms_lines_by_inv[oid].append(row_dict)

    def parse_int(val):
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return None

    def get_remarks(qty_alloc, wh_stock):
        """
        Priority: qty_allocated from WMS > wh_stock from inventory.
        - qty_alloc >= 1                    → FULFILL
        - qty_alloc == 0, wh_stock == 0     → CANCEL
        - qty_alloc == 0, wh_stock > 0      → FULFILL
        - qty_alloc is None, wh_stock == 0  → CANCEL  (not allocated, no stock)
        - qty_alloc is None, wh_stock > 0   → FULFILL (stock available)
        - both None                         → blank
        """
        if qty_alloc is not None:
            if qty_alloc >= 1:
                return "FULFILL"
            # qty_alloc == 0
            if wh_stock is not None:
                return "CANCEL" if wh_stock == 0 else "FULFILL"
        else:
            # qty_alloc is None — fall back to wh_stock only
            if wh_stock is not None:
                return "CANCEL" if wh_stock == 0 else "FULFILL"
        return ""

    # Deduplicate to one TC row per invoice — SKU detail comes from WMS lines anyway
    df_partial_dedup = df_partial.drop_duplicates(subset=["invoice_number"])

    sheet_row = 2
    for i, (_, r) in enumerate(df_partial_dedup.iterrows(), 1):
        inv_num   = str(r["invoice_number"]).strip()
        rf        = row_fill(r["nickname"], i)
        wms_val   = r["WMS"]
        wms_lines = wms_lines_by_inv.get(inv_num, [])

        if wms_lines:
            for line_idx, wd in enumerate(wms_lines):
                wms_sku     = str(wd.get("client_code_item", "")).strip()
                item_descr  = str(wd.get("item_descr", "")).strip()
                qty_ordered = parse_int(wd.get("quantity", ""))
                qty_alloc   = parse_int(wd.get("quantity_allocated", ""))
                raw_inv_val = inv_map.get(wms_sku)
                wh_stock    = parse_int(raw_inv_val) if raw_inv_val is not None else None
                remarks     = get_remarks(qty_alloc, wh_stock)

                is_first = (line_idx == 0)

                # Order-level cols (Invoice Number … WMS) — 7 cols, blank on sub-rows
                if is_first:
                    order_vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                                  r["order_item_status"], r["ordered_date"],
                                  r["nickname"], r["MP SLA"], r["WMS"]]
                else:
                    order_vals = ["", "", "", "", "", "", "", ""]

                vals = order_vals + [
                    wms_sku, item_descr,
                    qty_ordered if qty_ordered is not None else "",
                    qty_alloc   if qty_alloc   is not None else 0,
                    wh_stock    if wh_stock    is not None else "",
                    remarks,
                ]

                for ci, v in enumerate(vals, 1):
                    c = ws4.cell(sheet_row, ci, v)
                    c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
                    if not is_first and ci <= PARTIAL_ORDER_COLS:
                        # Sub-row order-level cells → white, no value
                        c.fill = WHITE_FILL
                    elif ci == PARTIAL_WMS_COL and is_first and wms_val in WMS_FILL:
                        c.fill = WMS_FILL[wms_val]
                    elif ci == len(PARTIAL_COLS):  # Remarks col
                        c.fill = (RED_FILL   if remarks == "CANCEL"  else
                                  GREEN_FILL if remarks == "FULFILL" else rf)
                    else:
                        c.fill = rf
                sheet_row += 1

        else:
            # No WMS lines — fall back to inventory stock
            stock    = r["_stock"]
            remarks  = get_remarks(None, stock if stock >= 0 else None)
            raw_sku  = str(r["custom_sku"]).strip()
            sku_base = raw_sku.split("x")[0] if raw_sku else ""
            wms_sku  = f"GSK_{sku_base}" if sku_base else ""

            vals = [r["order_number"], r["invoice_number"], r["payment_status"],
                    r["order_item_status"], r["ordered_date"],
                    r["nickname"], r["MP SLA"], r["WMS"],
                    wms_sku, "", "", "",
                    stock if stock >= 0 else "", remarks]

            for ci, v in enumerate(vals, 1):
                c = ws4.cell(sheet_row, ci, v)
                c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
                if ci == PARTIAL_WMS_COL and wms_val in WMS_FILL:
                    c.fill = WMS_FILL[wms_val]
                elif ci == len(PARTIAL_COLS):
                    c.fill = (RED_FILL   if remarks == "CANCEL"  else
                              GREEN_FILL if remarks == "FULFILL" else rf)
                else:
                    c.fill = rf
            sheet_row += 1

    # ── Sheet 5: ORIGINAL (All Data) ──────────────────────────────────────────
    ws6 = wb.create_sheet("ORIGINAL (All Data)")
    ws6.sheet_properties.tabColor = "1F1F1F"
    write_header_row(ws6, list(tc.columns))
    for ri, (_, r) in enumerate(tc.iterrows(), 2):
        for ci, col in enumerate(tc.columns, 1):
            c = ws6.cell(ri, ci, r[col])
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER

    # ── Sheet 7: RMA ──────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("RMA")
    ws7.sheet_properties.tabColor = "C00000"
    write_header_row(ws7, RMA_COLS, RMA_WIDTHS)
    for i, (_, r) in enumerate(df_rma.iterrows(), 1):
        vals = [r["order_number"], r["invoice_number"], r["order_status"],
                r["payment_status"], r["ordered_date"], r["nickname"], r["WMS"]]
        write_data_row(ws7, i + 1, vals, row_fill(r["nickname"], i),
                       wms_col_idx=7, wms_val=r["WMS"])

    # ── Sheet 8: NOT IN TC (MP orders missing from TC report) ────────────────
    NOT_IN_TC_COLS   = ["Order Number", "Marketplace", "Store", "MP Status",
                        "Ordered Date", "MP SLA", "Item Name", "Seller SKU"]
    NOT_IN_TC_WIDTHS = [22, 14, 20, 16, 22, 14, 50, 22]

    # Build set of all TC order IDs for fast lookup
    tc_order_ids = (
        set(tc["order_id"].dropna().astype(str).str.strip()) |
        set(tc["order_number"].dropna().astype(str).str.strip())
    )

    mp_missing_rows = []   # list of dicts

    # ── Shopee (reuse already-loaded `sh` DataFrame) ─────────────────────────
    SHOPEE_PENDING_KEYWORDS = ("order received", "to ship", "shipping")
    if "Order Status" in sh.columns:
        sh_status_lc = sh["Order Status"].str.lower()
        sh_pending_mask = sh_status_lc.str.contains("|".join(SHOPEE_PENDING_KEYWORDS), na=False)
        sh_pending = sh[sh_pending_mask]
        sh_missing = sh_pending[~sh_pending["Order ID"].isin(tc_order_ids)].drop_duplicates("Order ID")
        for _, sr in sh_missing.iterrows():
            mp_missing_rows.append({
                "Order Number": sr.get("Order ID", ""),
                "Marketplace":  "Shopee",
                "Store":        "shopee-SG",
                "MP Status":    sr.get("Order Status", ""),
                "Ordered Date": sr.get("Order Creation Date", ""),
                "MP SLA":       sr.get("Estimated Ship Out Date", ""),
                "Item Name":    sr.get("Product Name", ""),
                "Seller SKU":   sr.get("SKU Reference No.", ""),
            })

    # ── TikTok files ──────────────────────────────────────────────────────────
    TIKTOK_PENDING = {"to ship", "ready to ship", "awaiting shipment", "processing", "on hold"}
    for f in tiktok_files:
        try:
            raw_bytes  = f.read() if hasattr(f, "read") else f
            store_name = getattr(f, "name", "TikTok").replace(".xlsx", "")
            # Row 0 = headers, Row 1 = descriptions (skip), Row 2+ = data
            tt = pd.read_excel(BytesIO(raw_bytes), header=0, dtype=str, skiprows=[1])
            tt.columns = tt.columns.str.strip()
            tt["Order ID"]     = tt["Order ID"].astype(str).str.strip()
            tt["Order Status"] = tt["Order Status"].astype(str).str.strip()
            tt_pending = tt[tt["Order Status"].str.lower().isin(TIKTOK_PENDING)]
            tt_missing = tt_pending[~tt_pending["Order ID"].isin(tc_order_ids)].drop_duplicates("Order ID")
            for _, tr in tt_missing.iterrows():
                oid = str(tr.get("Order ID", "")).strip()
                if not oid or oid == "nan":
                    continue
                mp_status    = str(tr.get("Order Status", "")).strip()
                ordered_date = str(tr.get("Created Time", "")).strip()
                mp_sla       = add_one_day(ordered_date)
                mp_missing_rows.append({
                    "Order Number": oid,
                    "Marketplace":  "TikTok",
                    "Store":        store_name,
                    "MP Status":    mp_status,
                    "Ordered Date": ordered_date,
                    "MP SLA":       mp_sla,
                    "Item Name":    str(tr.get("Product Name", "")).strip(),
                    "Seller SKU":   str(tr.get("Seller SKU", "")).strip(),
                })
        except Exception as e:
            st.warning(f"Could not read TikTok file: {e}")

    # ── Lazada file ───────────────────────────────────────────────────────────
    LAZADA_PENDING = {"ready_to_ship", "confirmed", "packed"}
    if lazada_file is not None:
        try:
            laz_bytes = lazada_file.read() if hasattr(lazada_file, "read") else lazada_file
            laz = pd.read_excel(BytesIO(laz_bytes), dtype=str)
            laz.columns = laz.columns.str.strip()
            laz["orderNumber"] = laz["orderNumber"].astype(str).str.strip()
            laz["status"]      = laz["status"].astype(str).str.strip().str.lower()
            laz_pending = laz[laz["status"].isin(LAZADA_PENDING)]
            laz_missing = laz_pending[~laz_pending["orderNumber"].isin(tc_order_ids)]
            # Deduplicate by orderNumber (multiple items per order)
            laz_missing_dedup = laz_missing.drop_duplicates(subset=["orderNumber"])
            for _, lr in laz_missing_dedup.iterrows():
                ordered_date = str(lr.get("createTime", "")).strip()
                mp_sla       = add_one_day(ordered_date)
                mp_missing_rows.append({
                    "Order Number": lr.get("orderNumber", ""),
                    "Marketplace":  "Lazada",
                    "Store":        "lazada-SG",
                    "MP Status":    lr.get("status", ""),
                    "Ordered Date": ordered_date,
                    "MP SLA":       mp_sla,
                    "Item Name":    lr.get("itemName", ""),
                    "Seller SKU":   lr.get("sellerSku", ""),
                })
        except Exception as e:
            st.warning(f"Could not read Lazada file: {e}")

    ws8 = wb.create_sheet("NOT IN TC")
    ws8.sheet_properties.tabColor = "FF6600"
    write_header_row(ws8, NOT_IN_TC_COLS, NOT_IN_TC_WIDTHS)

    MP_NICK_FILL = {
        "Shopee": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "TikTok": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "Lazada": PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
    }
    MP_NICK_FILL_ALT = {
        "Shopee": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "TikTok": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "Lazada": PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
    }

    for i, mrow in enumerate(mp_missing_rows, 1):
        mp  = mrow["Marketplace"]
        rf  = MP_NICK_FILL_ALT[mp] if i % 2 == 0 else MP_NICK_FILL.get(mp, PatternFill())
        vals = [mrow[c] for c in NOT_IN_TC_COLS]
        for ci, v in enumerate(vals, 1):
            c = ws8.cell(i + 1, ci, v)
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER; c.fill = rf

    summary = {
        "FILTERED DATA":           len(df_filtered),
        "NOT PUSHED TO WMS":       len(df_not_pushed),
        "PARTIAL & NON ALLOCATED": len(df_partial),
        "ORIGINAL":                len(tc),
        "RMA":                     len(df_rma),
        "NOT IN TC":               len(mp_missing_rows),
    }
    return wb, summary


# ══════════════════════════════════════════════════════════════════════════════
#  NOT-PUSHED UPDATER
# ══════════════════════════════════════════════════════════════════════════════

def update_not_pushed(report_bytes, invoice_list, wms_status="PENDING_FULFILLMENT"):
    """
    Given an existing report (bytes) and a list of invoice numbers that have
    now been pushed to WMS:
      1. Read those rows from NOT PUSHED
      2. Set their WMS column to wms_status
      3. Append them to FILTERED DATA
      4. Remove them from NOT PUSHED and CLOSE & NOT FOUND
      5. Rebuild PIVOT
      PARTIAL, ORIGINAL, RMA are left completely untouched.

    Returns (updated_bytes, matched_count, not_found_set)
    """
    wb        = load_workbook(BytesIO(report_bytes))
    today_str = datetime.today().strftime("%Y-%m-%d")
    invoices  = {str(i).strip().upper() for i in invoice_list if str(i).strip()}

    NUM_COLS    = 8   # all standard sheets have exactly 8 data columns
    WMS_COL    = 7   # 0-based index of WMS in an 8-element row
    NICK_COL   = 5   # 0-based index of Nickname

    def read_rows(ws):
        """Return (headers_list, data_rows_list) reading exactly NUM_COLS columns."""
        headers = [ws.cell(1, c).value for c in range(1, NUM_COLS + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            rv = [ws.cell(r, c).value for c in range(1, NUM_COLS + 1)]
            if any(v is not None for v in rv):
                rows.append(rv)
        return headers, rows

    def rewrite_sheet(ws, rows):
        """
        Clear all data rows (with a safe ceiling) and write new rows.
        Header row (row 1) is preserved.
        """
        ceiling = max(ws.max_row or 1, len(rows) + 20)
        for r in range(2, ceiling + 1):
            for c in range(1, NUM_COLS + 1):
                cell = ws.cell(r, c)
                cell.value  = None
                cell.fill   = PatternFill()
                cell.font   = BODY_FONT
                cell.border = NO_BORDER
        for i, rv in enumerate(rows, 1):
            nick    = rv[NICK_COL] if len(rv) > NICK_COL else ""
            wms_val = rv[WMS_COL]  if len(rv) > WMS_COL  else None
            rf      = row_fill(nick, i)
            write_data_row(ws, i + 1, rv, rf, wms_col_idx=8, wms_val=wms_val)

    # 1. Read FILTERED DATA
    ws_fd              = wb["FILTERED DATA"]
    fd_headers, fd_rows = read_rows(ws_fd)

    # 2. Partition NOT PUSHED TO WMS into matched vs kept
    ws_np              = wb["NOT PUSHED TO WMS"]
    np_headers, np_rows = read_rows(ws_np)
    inv_col            = np_headers.index("Invoice Number")

    matched, kept_np, matched_inv = [], [], set()
    for rv in np_rows:
        key = str(rv[inv_col] or "").strip().upper()
        if key in invoices:
            updated = list(rv[:WMS_COL]) + [wms_status]  # replace WMS column
            matched.append(updated)
            matched_inv.add(key)
        else:
            kept_np.append(rv)

    # 3. Merge matched into FILTERED DATA
    final_fd = fd_rows + matched

    # 4. Rewrite the two changed sheets (PARTIAL, ORIGINAL, RMA, NOT IN TC untouched)
    rewrite_sheet(ws_fd,  final_fd)
    rewrite_sheet(ws_np,  kept_np)

    # 5. Rebuild PIVOT only — PARTIAL, ORIGINAL, RMA untouched
    build_pivot_sheet(wb["PIVOT"], final_fd, fd_headers, today_str)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), len(matched), invoices - matched_inv



# ══════════════════════════════════════════════════════════════════════════════
#  NOT-IN-TC UPDATER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def update_not_in_tc_manual(report_bytes, new_row):
    wb        = load_workbook(BytesIO(report_bytes))
    today_str = datetime.today().strftime("%Y-%m-%d")
    NUM_COLS  = 8
    NICK_COL  = 5
    WMS_COL   = 7

    # Read FILTERED DATA
    ws_fd = wb["FILTERED DATA"]
    fd_headers = [ws_fd.cell(1, c).value for c in range(1, NUM_COLS + 1)]
    fd_rows = []
    for r in range(2, ws_fd.max_row + 1):
        rv = [ws_fd.cell(r, c).value for c in range(1, NUM_COLS + 1)]
        if any(v is not None for v in rv):
            fd_rows.append(rv)

    # Append new row
    new_rv = [
        new_row.get("Order Number", ""),
        new_row.get("Invoice Number", ""),
        new_row.get("Payment Status", ""),
        new_row.get("Order Item Status", ""),
        new_row.get("Ordered Date", ""),
        new_row.get("Nickname", ""),
        new_row.get("MP SLA", ""),
        new_row.get("WMS", ""),
    ]
    final_fd = fd_rows + [new_rv]

    # Rewrite FILTERED DATA
    ceiling = max(ws_fd.max_row or 1, len(final_fd) + 20)
    for r in range(2, ceiling + 1):
        for c in range(1, NUM_COLS + 1):
            cell = ws_fd.cell(r, c)
            cell.value = None; cell.fill = PatternFill()
            cell.font = BODY_FONT; cell.border = NO_BORDER
    for i, rv in enumerate(final_fd, 1):
        nick    = rv[NICK_COL] if len(rv) > NICK_COL else ""
        wms_val = rv[WMS_COL]  if len(rv) > WMS_COL  else None
        write_data_row(ws_fd, i + 1, rv, row_fill(nick, i), wms_col_idx=8, wms_val=wms_val)

    # Remove order from NOT IN TC
    order_num = str(new_row.get("Order Number", "")).strip().upper()
    _rewrite_not_in_tc(wb, order_nums_to_remove={order_num})

    # Rebuild PIVOT
    build_pivot_sheet(wb["PIVOT"], final_fd, fd_headers, today_str)

    out = BytesIO(); wb.save(out); out.seek(0)
    return out.read()


def refresh_not_in_tc(report_bytes, new_tc_file, new_wms_files):
    wb        = load_workbook(BytesIO(report_bytes))
    today_str = datetime.today().strftime("%Y-%m-%d")
    NUM_COLS  = 8
    NICK_COL  = 5
    WMS_COL   = 7

    # Load fresh TC
    tc_new = pd.read_csv(BytesIO(new_tc_file.read()), dtype=str)
    tc_new = tc_new.apply(lambda col: col.str.replace("\r", "", regex=False).str.strip())
    tc_new.replace("nan", "", inplace=True)
    new_tc_ids = (
        set(tc_new["order_id"].dropna().astype(str).str.strip()) |
        set(tc_new["order_number"].dropna().astype(str).str.strip())
    )

    # Load fresh WMS
    wms_dfs = []
    for f in new_wms_files:
        try:
            wms_dfs.append(pd.read_excel(BytesIO(f.read()), dtype=str))
        except: pass
    new_wms_map = {}
    if wms_dfs:
        wms_new = pd.concat(wms_dfs, ignore_index=True)
        wms_new.columns = wms_new.columns.str.strip()
        wms_new["client_order_id"] = wms_new["client_order_id"].astype(str).str.strip()
        wms_new["status_so"]       = wms_new["status_so"].astype(str).str.strip()
        new_wms_map = wms_new.drop_duplicates("client_order_id").set_index("client_order_id")["status_so"].to_dict()

    if "NOT IN TC" not in wb.sheetnames:
        out = BytesIO(); wb.save(out); out.seek(0)
        return out.read(), 0

    # Read NOT IN TC
    ws_nitc = wb["NOT IN TC"]
    nitc_headers = [ws_nitc.cell(1, c).value for c in range(1, ws_nitc.max_column + 1)]
    on_col = nitc_headers.index("Order Number") if "Order Number" in nitc_headers else 0

    resolved_order_nums = set()
    resolved_tc_rows    = []
    for r in range(2, ws_nitc.max_row + 1):
        rv = [ws_nitc.cell(r, c).value for c in range(1, ws_nitc.max_column + 1)]
        if any(v is not None for v in rv):
            order_num = str(rv[on_col] or "").strip()
            if order_num in new_tc_ids:
                resolved_order_nums.add(order_num.upper())
                resolved_tc_rows.append(order_num)

    # Read and extend FILTERED DATA
    ws_fd = wb["FILTERED DATA"]
    fd_headers = [ws_fd.cell(1, c).value for c in range(1, NUM_COLS + 1)]
    fd_rows = []
    for r in range(2, ws_fd.max_row + 1):
        rv = [ws_fd.cell(r, c).value for c in range(1, NUM_COLS + 1)]
        if any(v is not None for v in rv):
            fd_rows.append(rv)

    for order_num in resolved_tc_rows:
        tc_match = tc_new[(tc_new["order_id"] == order_num) | (tc_new["order_number"] == order_num)]
        if len(tc_match):
            r = tc_match.iloc[0]
            inv_num    = r.get("invoice_number", "")
            wms_status = new_wms_map.get(inv_num, "NOT FOUND") if inv_num else "NOT FOUND"
            fd_rows.append([
                r.get("order_number", order_num),
                inv_num,
                r.get("payment_status", ""),
                r.get("order_item_status", ""),
                r.get("ordered_date", ""),
                r.get("nickname", ""),
                "",
                wms_status,
            ])

    # Rewrite FILTERED DATA
    ceiling = max(ws_fd.max_row or 1, len(fd_rows) + 20)
    for r in range(2, ceiling + 1):
        for c in range(1, NUM_COLS + 1):
            cell = ws_fd.cell(r, c)
            cell.value = None; cell.fill = PatternFill()
            cell.font = BODY_FONT; cell.border = NO_BORDER
    for i, rv in enumerate(fd_rows, 1):
        nick    = rv[NICK_COL] if len(rv) > NICK_COL else ""
        wms_val = rv[WMS_COL]  if len(rv) > WMS_COL  else None
        write_data_row(ws_fd, i + 1, rv, row_fill(nick, i), wms_col_idx=8, wms_val=wms_val)

    # Remove resolved orders from NOT IN TC
    _rewrite_not_in_tc(wb, order_nums_to_remove=resolved_order_nums)

    # Rebuild PIVOT
    build_pivot_sheet(wb["PIVOT"], fd_rows, fd_headers, today_str)

    out = BytesIO(); wb.save(out); out.seek(0)
    return out.read(), len(resolved_order_nums)


def _rewrite_not_in_tc(wb, order_nums_to_remove):
    """Helper: remove rows from NOT IN TC sheet where Order Number is in the given set."""
    if "NOT IN TC" not in wb.sheetnames:
        return
    ws_nitc = wb["NOT IN TC"]
    nitc_headers = [ws_nitc.cell(1, c).value for c in range(1, ws_nitc.max_column + 1)]
    on_col = nitc_headers.index("Order Number") if "Order Number" in nitc_headers else 0
    mp_col = nitc_headers.index("Marketplace")  if "Marketplace"  in nitc_headers else 1
    num_cols = len(nitc_headers)

    kept = []
    for r in range(2, ws_nitc.max_row + 1):
        rv = [ws_nitc.cell(r, c).value for c in range(1, num_cols + 1)]
        if any(v is not None for v in rv):
            if str(rv[on_col] or "").strip().upper() not in order_nums_to_remove:
                kept.append(rv)

    MP_FILL = {
        "Shopee": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "TikTok": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "Lazada": PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
    }
    MP_FILL_ALT = {
        "Shopee": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "TikTok": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "Lazada": PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
    }
    ceiling = max(ws_nitc.max_row or 1, len(kept) + 20)
    for r in range(2, ceiling + 1):
        for c in range(1, num_cols + 1):
            cell = ws_nitc.cell(r, c)
            cell.value = None; cell.fill = PatternFill()
            cell.font = BODY_FONT; cell.border = NO_BORDER
    for i, rv in enumerate(kept, 1):
        mp = str(rv[mp_col] or "")
        rf = MP_FILL_ALT.get(mp, PatternFill()) if i % 2 == 0 else MP_FILL.get(mp, PatternFill())
        for ci, v in enumerate(rv, 1):
            c = ws_nitc.cell(i + 1, ci, v)
            c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER; c.fill = rf



# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI  —  single-tab, inline update design
# ══════════════════════════════════════════════════════════════════════════════

st.title("📦 GSK Pending Order Report")
st.markdown("---")

# ── Session-state initialisation ─────────────────────────────────────────────
for _k, _v in {
    "report_bytes":    None,
    "report_filename": None,
    "mp_data":         None,   # dict: {order_id -> row_dict} from all MP files
    "np_wms_rows":     None,   # list of dicts for NOT PUSHED TO WMS preview
    "nitc_rows":       None,   # list of dicts for NOT IN TC preview
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


def _read_mp_lookup(shopee_bytes, tiktok_bytes_list, lazada_bytes):
    """
    Build a lookup dict  order_id -> {payment_status, order_item_status,
    ordered_date, nickname, mp_sla, mp_status, item_name, seller_sku}
    from all MP files uploaded during report generation.

    Shopee  : Order Status (col C), Estimated Ship Out Date (col I) = MP SLA,
              Order Creation Date (col K) = ordered date
    Lazada  : status (col BO) = MP status, createTime = ordered date,
              MP SLA = ordered date + 1 day
    TikTok  : Order ID only — no extra data available in current export format
    """
    lookup = {}

    # ── Shopee ────────────────────────────────────────────────────────────────
    if shopee_bytes:
        try:
            sh = pd.read_excel(BytesIO(shopee_bytes), dtype=str)
            sh.columns = sh.columns.str.strip()
            for _, r in sh.iterrows():
                oid = str(r.get("Order ID", "")).strip()
                if not oid or oid == "nan":
                    continue
                mp_status  = str(r.get("Order Status", "")).strip()
                pay_status = "PENDING" if "unpaid" in mp_status.lower() else "COMPLETED"
                sla        = str(r.get("Estimated Ship Out Date", "")).strip()
                sla_clean  = sla[:10] if sla and sla != "nan" else ""
                lookup[oid] = {
                    "payment_status":    pay_status,
                    "order_item_status": "READY TO SHIP",
                    "ordered_date":      str(r.get("Order Creation Date", "")).strip(),
                    "nickname":          "shopee-SG",
                    "mp_sla":            sla_clean,
                    "mp_status":         mp_status,
                    "item_name":         str(r.get("Product Name", "")).strip(),
                    "seller_sku":        str(r.get("SKU Reference No.", "")).strip(),
                }
        except Exception:
            pass

    # ── TikTok ────────────────────────────────────────────────────────────────
    # Row 0 = column headers, Row 1 = descriptions (skip), Row 2+ = data
    for tt_bytes, store in (tiktok_bytes_list or []):
        try:
            tt = pd.read_excel(BytesIO(tt_bytes), header=0, dtype=str, skiprows=[1])
            tt.columns = tt.columns.str.strip()
            for _, r in tt.iterrows():
                oid = str(r.get("Order ID", "")).strip()
                if not oid or oid == "nan":
                    continue
                mp_status    = str(r.get("Order Status", "")).strip()
                pay_status   = "PENDING" if mp_status.lower() == "unpaid" else "COMPLETED"
                ordered_date = str(r.get("Created Time", "")).strip()
                mp_sla       = add_one_day(ordered_date)
                lookup[oid] = {
                    "payment_status":    pay_status,
                    "order_item_status": "READY TO SHIP",
                    "ordered_date":      ordered_date,
                    "nickname":          store,
                    "mp_sla":            mp_sla,
                    "mp_status":         mp_status,
                    "item_name":         str(r.get("Product Name", "")).strip(),
                    "seller_sku":        str(r.get("Seller SKU", "")).strip(),
                }
        except Exception:
            pass

    # ── Lazada ────────────────────────────────────────────────────────────────
    if lazada_bytes:
        try:
            laz = pd.read_excel(BytesIO(lazada_bytes), dtype=str)
            laz.columns = laz.columns.str.strip()
            for _, r in laz.iterrows():
                oid = str(r.get("orderNumber", "")).strip()
                if not oid or oid == "nan":
                    continue
                mp_status    = str(r.get("status", "")).strip()
                pay_status   = "PENDING" if "unpaid" in mp_status.lower() else "COMPLETED"
                ordered_date = str(r.get("createTime", "")).strip()
                mp_sla       = add_one_day(ordered_date)
                if oid not in lookup:
                    lookup[oid] = {
                        "payment_status":    pay_status,
                        "order_item_status": "READY TO SHIP",
                        "ordered_date":      ordered_date,
                        "nickname":          "lazada-SG",
                        "mp_sla":            mp_sla,
                        "mp_status":         mp_status,
                        "item_name":         str(r.get("itemName", "")).strip(),
                        "seller_sku":        str(r.get("sellerSku", "")).strip(),
                    }
        except Exception:
            pass

    return lookup


def _load_preview_from_report(report_bytes):
    """Read NOT PUSHED TO WMS and NOT IN TC rows from the report bytes."""
    wb = load_workbook(BytesIO(report_bytes), read_only=True)
    np_rows, nitc_rows = [], []

    if "NOT PUSHED TO WMS" in wb.sheetnames:
        ws = wb["NOT PUSHED TO WMS"]
        hdrs = [ws.cell(1, c).value for c in range(1, 9)]
        for r in range(2, ws.max_row + 1):
            rv = [ws.cell(r, c).value for c in range(1, 9)]
            if any(v is not None for v in rv):
                np_rows.append(dict(zip(hdrs, rv)))

    if "NOT IN TC" in wb.sheetnames:
        ws2 = wb["NOT IN TC"]
        hdrs2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
        for r in range(2, ws2.max_row + 1):
            rv2 = [ws2.cell(r, c).value for c in range(1, ws2.max_column + 1)]
            if any(v is not None for v in rv2):
                nitc_rows.append(dict(zip(hdrs2, rv2)))

    wb.close()
    return np_rows, nitc_rows


# ── File uploaders ────────────────────────────────────────────────────────────
st.subheader("Upload daily source files")

col_a, col_b = st.columns(2)
with col_a:
    tc_file     = st.file_uploader("TC Order Report (.csv)",    type=["csv"],  key="tc")
    shopee_file = st.file_uploader("Shopee Order File (.xlsx)", type=["xlsx"], key="shopee")
with col_b:
    wms_files   = st.file_uploader("WMS File(s) (.xlsx)",       type=["xlsx"], key="wms",
                                    accept_multiple_files=True)
    inv_file    = st.file_uploader("Inventory Report (.xlsx)",  type=["xlsx"], key="inv")

st.markdown("##### MP Order Files *(optional — for NOT IN TC check)*")
col_c, col_d = st.columns(2)
with col_c:
    tiktok_files = st.file_uploader("TikTok Order File(s) (.xlsx) — both stores",
                                     type=["xlsx"], key="tiktok", accept_multiple_files=True)
with col_d:
    lazada_file = st.file_uploader("Lazada Order File (.xlsx)", type=["xlsx"], key="lazada")


# ── Generate Report button ────────────────────────────────────────────────────
if st.button("🚀 Generate Report", type="primary", use_container_width=True):
    if not all([tc_file, shopee_file, wms_files, inv_file]):
        st.error("Please upload all four required files before generating.")
    else:
        with st.spinner("Processing files…"):
            try:
                # Read all bytes immediately (UploadedFile exhausts after first read)
                shopee_bytes = shopee_file.read()
                tiktok_bytes_list = [(f.read(), f.name.replace(".xlsx",""))
                                     for f in (tiktok_files or [])]
                lazada_bytes = lazada_file.read() if lazada_file else None

                # Wrap back as file-like for generate_report
                import io
                class _Wrap:
                    def __init__(self, b, name=""): self._b = b; self.name = name
                    def read(self): return self._b

                wb, summary = generate_report(
                    tc_file,
                    _Wrap(shopee_bytes, "shopee"),
                    [_Wrap(f.read() if hasattr(f,"read") else f) for f in wms_files],
                    inv_file,
                    tiktok_files=[_Wrap(b, n) for b, n in tiktok_bytes_list],
                    lazada_file=_Wrap(lazada_bytes) if lazada_bytes else None,
                )
                buf = BytesIO(); wb.save(buf)
                report_bytes = buf.getvalue()
                fname = f"GSK_Pending_Order_Report_{datetime.today().strftime('%Y%m%d')}.xlsx"

                # Store in session
                st.session_state["report_bytes"]    = report_bytes
                st.session_state["report_filename"] = fname
                st.session_state["mp_data"]         = _read_mp_lookup(
                    shopee_bytes, tiktok_bytes_list, lazada_bytes)
                np_rows, nitc_rows = _load_preview_from_report(report_bytes)
                st.session_state["np_wms_rows"] = np_rows
                st.session_state["nitc_rows"]   = nitc_rows

                st.success("✅ Report generated!")

                # Summary metrics
                icons = {"FILTERED DATA": "🔵", "NOT PUSHED TO WMS": "🔴",
                         "PARTIAL & NON ALLOCATED": "🟠",
                         "ORIGINAL": "⬛", "RMA": "🔴", "NOT IN TC": "🟡"}
                mcols = st.columns(len(summary))
                for col, (k, v) in zip(mcols, summary.items()):
                    col.metric(f"{icons.get(k,'')} {k}", v)

                st.download_button("⬇️ Download Report", data=report_bytes,
                                    file_name=fname, use_container_width=True,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error(f"Error generating report: {e}"); st.exception(e)


# ── Only show previews + update UI if report exists ──────────────────────────
if st.session_state.get("report_bytes"):

    # ══════════════════════════════════════════════════════════════════════════
    #  SECTION 1 — NOT PUSHED TO WMS
    # ══════════════════════════════════════════════════════════════════════════
    np_rows = st.session_state.get("np_wms_rows") or []
    if np_rows:
        st.markdown("---")
        st.markdown(f"### 🔴 Not Pushed to WMS — {len(np_rows)} order(s)")
        st.caption("Select WMS status and click Update for each order that has now been pushed.")

        for idx, row in enumerate(np_rows):
            with st.expander(
                f"📋 {row.get('Order Number','—')}  |  {row.get('Invoice Number','—')}  "
                f"|  {row.get('Nickname','—')}  |  SLA: {row.get('MP SLA','—')}",
                expanded=False,
            ):
                c1, c2, c3 = st.columns([2, 2, 1])
                with c1:
                    st.markdown(f"**Invoice:** `{row.get('Invoice Number','')}`")
                    st.markdown(f"**Status:** {row.get('Order Item Status','')}")
                    st.markdown(f"**Ordered:** {row.get('Ordered Date','')}")
                with c2:
                    new_wms = st.selectbox(
                        "WMS Status",
                        ["PENDING_FULFILLMENT", "FULFILLED", "PARTIAL_ALLOCATED"],
                        key=f"np_wms_{idx}",
                    )
                with c3:
                    st.markdown("<br>", unsafe_allow_html=True)
                    do_update = st.button("✅ Update", key=f"np_btn_{idx}", use_container_width=True)

                if do_update:
                    inv_num = row.get("Invoice Number", "")
                    with st.spinner(f"Updating {inv_num}…"):
                        try:
                            updated_bytes, matched, _ = update_not_pushed(
                                st.session_state["report_bytes"], [inv_num], new_wms)
                            st.session_state["report_bytes"] = updated_bytes
                            np_rows2, nitc_rows2 = _load_preview_from_report(updated_bytes)
                            st.session_state["np_wms_rows"] = np_rows2
                            st.session_state["nitc_rows"]   = nitc_rows2
                            st.session_state[f"np_updated_{idx}"] = True
                        except Exception as e:
                            st.error(f"Error: {e}")

                if st.session_state.get(f"np_updated_{idx}"):
                    inv_num = row.get("Invoice Number", "")
                    st.success(f"✅ `{inv_num}` moved to FILTERED DATA. PIVOT rebuilt.")
                    st.download_button(
                        "⬇️ Download Updated Report",
                        data=st.session_state["report_bytes"],
                        file_name=st.session_state["report_filename"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_np_{idx}",
                        use_container_width=True,
                    )
    else:
        if st.session_state.get("report_bytes"):
            st.markdown("---")
            st.success("✅ All orders have been pushed to WMS — nothing pending.")

    # ══════════════════════════════════════════════════════════════════════════
    #  SECTION 2 — NOT IN TC
    # ══════════════════════════════════════════════════════════════════════════
    nitc_rows = st.session_state.get("nitc_rows") or []
    if nitc_rows:
        st.markdown("---")
        st.markdown(f"### 🟡 Not in TC — {len(nitc_rows)} order(s)")
        st.caption("Enter Invoice Number + WMS status. Other details will be pulled from the MP files automatically.")

        mp_lookup = st.session_state.get("mp_data") or {}

        for idx, row in enumerate(nitc_rows):
            order_num  = row.get("Order Number", "")
            mp_info    = mp_lookup.get(order_num, {})
            _item_raw  = row.get("Item Name") or row.get("item_name") or mp_info.get("item_name") or "—"
            item_label = str(_item_raw)[:40]

            # Pull display values — prefer sheet data, fall back to mp_lookup
            disp_ordered  = row.get("Ordered Date") or mp_info.get("ordered_date") or "—"
            disp_sla      = row.get("MP SLA")      or mp_info.get("mp_sla")       or "—"
            disp_status   = row.get("MP Status")   or "—"
            disp_nickname = mp_info.get("nickname") or row.get("Store") or "—"
            disp_item     = str(row.get("Item Name") or mp_info.get("item_name") or "—")
            disp_sku      = str(row.get("Seller SKU") or mp_info.get("seller_sku") or "—")

            with st.expander(
                f"📋 {order_num}  |  {row.get('Marketplace','—')}  "
                f"|  {row.get('Store','—')}  |  {item_label}",
                expanded=False,
            ):
                c_info, c_inputs, c_btn = st.columns([2, 2, 1])
                with c_info:
                    st.markdown(f"**Order Number:** `{order_num}`")
                    st.markdown(f"**Marketplace:** {row.get('Marketplace','')}")
                    st.markdown(f"**MP Status:** {disp_status}")
                    st.markdown(f"**Ordered Date:** {disp_ordered}")
                    st.markdown(f"**MP SLA:** {disp_sla}")
                    st.markdown(f"**Nickname:** {disp_nickname}")
                    st.markdown(f"**Item:** {disp_item[:60]}")
                    st.markdown(f"**SKU:** {disp_sku}")
                with c_inputs:
                    inv_input = st.text_input("Invoice Number *",
                                              key=f"nitc_inv_{idx}",
                                              placeholder="e.g. GSK00364795")
                    wms_input = st.selectbox("WMS Status",
                                             ["PENDING_FULFILLMENT", "FULFILLED",
                                              "PARTIAL_ALLOCATED", "NOT FOUND"],
                                             key=f"nitc_wms_{idx}")
                with c_btn:
                    st.markdown("<br><br>", unsafe_allow_html=True)
                    do_nitc_update = st.button("✅ Update", key=f"nitc_btn_{idx}",
                                               use_container_width=True)

                if do_nitc_update:
                    if not inv_input.strip():
                        st.error("Invoice Number is required.")
                    else:
                        new_row = {
                            "Order Number":      order_num,
                            "Invoice Number":    inv_input.strip(),
                            "Payment Status":    mp_info.get("payment_status", "COMPLETED"),
                            "Order Item Status": mp_info.get("order_item_status", "READY TO SHIP"),
                            "Ordered Date":      mp_info.get("ordered_date") or str(row.get("Ordered Date", "")),
                            "Nickname":          mp_info.get("nickname") or row.get("Store", ""),
                            "MP SLA":            mp_info.get("mp_sla") or str(row.get("MP SLA", "")),
                            "WMS":               wms_input,
                        }
                        with st.spinner(f"Updating {order_num}…"):
                            try:
                                updated_bytes = update_not_in_tc_manual(
                                    st.session_state["report_bytes"], new_row)
                                st.session_state["report_bytes"] = updated_bytes
                                np_rows2, nitc_rows2 = _load_preview_from_report(updated_bytes)
                                st.session_state["np_wms_rows"] = np_rows2
                                st.session_state["nitc_rows"]   = nitc_rows2
                                st.session_state[f"nitc_updated_{idx}"] = {
                                    "order_num": order_num,
                                    "inv":       inv_input.strip(),
                                    "wms":       wms_input,
                                }
                            except Exception as e:
                                st.error(f"Error: {e}")

                if st.session_state.get(f"nitc_updated_{idx}"):
                    _u = st.session_state[f"nitc_updated_{idx}"]
                    st.success(
                        f"✅ `{_u['order_num']}` → FILTERED DATA  "
                        f"(Invoice: `{_u['inv']}`, WMS: `{_u['wms']}`). PIVOT rebuilt."
                    )
                    st.download_button(
                        "⬇️ Download Updated Report",
                        data=st.session_state["report_bytes"],
                        file_name=st.session_state["report_filename"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_nitc_{idx}",
                        use_container_width=True,
                    )
    else:
        if st.session_state.get("report_bytes"):
            st.markdown("---")
            st.success("✅ All MP orders are present in TC — nothing missing.")
